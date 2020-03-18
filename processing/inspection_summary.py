"""
検査実施サマリ.xlsx用
"""

from openpyxl import load_workbook
import os

FILENAME = "検査実施サマリ.xlsx"

def parse_inspection_summary():
    paths = [os.path.abspath(os.path.dirname(__file__)), '..', 'data', FILENAME]
    f = os.path.join(*paths)
    wb = load_workbook(f)
    ws = wb.active
    return ws['A2'].value

if __name__ == '__main__':
    print(parse_inspection_summary())