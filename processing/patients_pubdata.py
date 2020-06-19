"""
- https://www.pref.chiba.lg.jp/shippei/press/2019/documents/chiba_corona_data.xlsx
- 新型コロナウイルス感染者数（検査確定日、公表日、7日間平均）

上記を集計したものを出力

詳細なデータ
```
  "patients": {
    "date": "2020/06/16 15:21",
    "data": [
      {
        "日付": "1/30/2020",
        "確定数": 2,
        "公表数": 2,
        "小計": 2
      },
  },
```

"""

from openpyxl import load_workbook
import os
import glob
from datetime import timedelta
from functools import reduce
from pathlib import Path
import sys
sys.path.append(str(Path('__file__').resolve().parent))
from common import excel_date
modified_date = None

def _empty_data(date):
    return {
        "日付": date.strftime('%-m/%-d/%Y'),
        "確定数": 0,
        "公表数": 0,
        "小計": 0
    }

def patients_modified(filepath):
    wb = load_workbook(filepath)
    return wb.properties.modified

def _patients_dataset_from_chiba_pref(filepath):
    SHEETNAME = "新型コロナウイルス感染者数（検査確定日、公表日、7日間平均）"
    """
    感染者（データセット）の処理
    """
    data = {}
    wb = load_workbook(filepath)
    ws = wb[SHEETNAME]
    i = 0
    for row in ws.values:
        i += 1
        if i in [1, 2, 3, 4]:  # pass header
            continue
        if not row[1]:  # pass empty row
            continue
        definite_date = row[1]
        target_date = definite_date.date()
        if target_date not in data.keys():
            data[target_date] = _empty_data(target_date)
        data[target_date]["確定数"] = row[2]
        data[target_date]["小計"] = row[2]
        data[target_date]["公表数"] = row[5]
    return data

def _fill_data(data):
    """
    空の日付にデータを埋める
    """
    from_day = min(data.keys())
    to_day = max(data.keys())
    for i in range((to_day - from_day).days + 1):
        d = from_day + timedelta(i)
        if not d in data.keys():
            data[d] = _empty_data(d)
    return data

def _data_to_patients(data):
    """
    patientsのデータを作成
    """
    patients = []
    patients_summary_data = {
        "確定数": [],
        "公表数": []
    }
    patients_summary_labels = []
    keys = sorted(data.keys())
    for key in keys:
        patients.append(data[key])
        patients_summary_data["確定数"].append(data[key]["確定数"])
        patients_summary_data["公表数"].append(data[key]["公表数"])
        patients_summary_labels.append(key.strftime("%-m/%-d"))
    return patients, patients_summary_data, patients_summary_labels

def _total_count(data):
    total_conf_count = reduce(lambda x, y: x + y, [x["確定数"] for x in data.values()])
    total_pub_count = reduce(lambda x, y: x + y, [x["公表数"] for x in data.values()])
    return total_conf_count, total_pub_count

def parse_patients_per_date(filepath):
    data = _patients_dataset_from_chiba_pref(filepath)
    data = _fill_data(data)
    (total_conf_count, total_pub_count) = _total_count(data)
    return _data_to_patients(data), total_conf_count, total_pub_count

if __name__ == '__main__':
    filename = "chiba.xlsx"
    filepath = os.path.join(*[os.path.abspath(os.path.dirname(__file__)), '../data', filename])
    print(parse_patients_per_date(filepath))
