"""
- https://www.pref.chiba.lg.jp/shippei/press/2019/documents/chiba_corona_data.xlsx
- 新型コロナウイルス感染者数（累積、公表日別）
上記を集計したものを出力

"""

from openpyxl import load_workbook
import os
from pathlib import Path
import sys
sys.path.append(str(Path('__file__').resolve().parent))


def patients_modified(filepath):
    wb = load_workbook(filepath)
    return wb.properties.modified


def _patients_summary_dataset_from_chiba_pref(filepath):
    SHEETNAME = "新型コロナウイルス感染者数（累積、公表日別）"
    """
    感染者（データセット）の処理
    """
    wb = load_workbook(filepath)
    ws = wb[SHEETNAME]
    i = 0
    for row in ws.values:
        i += 1
        if i in [1, 2, 3, 4]:  # pass header
            continue
        if not row[1]:  # pass empty row
            continue
        last_index = i
    i = last_index
    target_date = ws.cell(i, 2).value.date().strftime('%-m/%-d/%Y')
    patients_count = ws.cell(i, 3).value
    hospital_count = ws.cell(i, 4).value
    hospital_waiting_count = ws.cell(i, 5).value
    hotel_stay_count = ws.cell(i, 6).value
    home_stay_count = ws.cell(i, 7).value
    discharge_count = ws.cell(i, 8).value
    finish_stay_count = ws.cell(i, 9).value
    death_count = ws.cell(i, 10).value
    other_count = ws.cell(i, 11).value
    severe_injury_count = ws.cell(i, 12).value
    return (target_date, patients_count, hospital_count, hospital_waiting_count, hotel_stay_count, home_stay_count, discharge_count, finish_stay_count, death_count, other_count, severe_injury_count)


def parse_patients_per_date(filepath):
    return _patients_summary_dataset_from_chiba_pref(filepath)


if __name__ == '__main__':
    filename = "chiba.xlsx"
    filepath = os.path.join(*[os.path.abspath(os.path.dirname(__file__)), '../data', filename])
    print(parse_patients_per_date(filepath))
