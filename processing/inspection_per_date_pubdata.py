"""
- https://www.pref.chiba.lg.jp/shippei/press/2019/documents/chiba_corona_data.xlsx
- PCR検査状況シート

上記を集計したものを出力

詳細なデータ
```
{
    "日付": "1/25/2020",
    "陽性": "0 ",
    "陰性": "0 ",
    "小計": "3 ",
},
```

"""
from openpyxl import load_workbook
import os
from datetime import timedelta
from datetime import datetime
from functools import reduce
from pathlib import Path
import sys
sys.path.append(str(Path('__file__').resolve().parent))
modified_date = None


def _empty_data(date):
    return {
        "日付": date.strftime('%-m/%-d/%Y'),
        "陽性": 0,
        "陰性": 0,
        "小計": 0
    }


def inspections_modified(filepath):
    wb = load_workbook(filepath)
    return wb.properties.modified


def _inspection_dataset_from_chiba_pref(filepath):
    SHEETNAME = "PCR検査状況"
    """
    検査実績（データセット）の処理
    """
    data = {}
    wb = load_workbook(filepath)
    ws = wb[SHEETNAME]
    i = 0
    for row in ws.values:
        i += 1
        if i == 2:  # modified date TBD will use on inspections_modified()
            modified_date = row[4]  # original data is  like "令和2年6月14日時点"
            continue
        if i in [1, 2, 3, 4]:  # pass header
            continue
        if not row[1]:  # pass empty row
            continue
        definite_date = row[1]
        # もし日付が文字列であれば
        if (type(definite_date) == type('1月1日')):
            month = definite_date.split('月')[0]
            day = definite_date.split('月')[1].split('日')[0]
            target_date = datetime(2020,int(month),int(day))
        else:
            target_date = definite_date.date()

        if target_date not in data.keys():
            data[target_date] = _empty_data(target_date)
        data[target_date]["陽性"] = row[2]
        data[target_date]["陰性"] = row[3]
        data[target_date]["小計"] = row[4]
    return data


def _fill_data(data):
    """
    空の日付にデータを埋める
    """
    from_day = min(data.keys())
    to_day = max(data.keys())
    for i in range((to_day - from_day).days + 1):
        d = from_day + timedelta(i)
        if d not in data.keys():
            data[d] = _empty_data(d)
    return data


def _data_to_inspections(data):
    """
    inspectionsのデータを作成
    """
    inspections = []
    inspections_summary_data = {
        "陽性": [],
        "陰性": []
    }
    inspections_summary_labels = []
    keys = sorted(data.keys())
    for key in keys:
        inspections.append(data[key])
        inspections_summary_data["陽性"].append(data[key]["陽性"])
        inspections_summary_data["陰性"].append(data[key]["陰性"])
        inspections_summary_labels.append(key.strftime("%-m/%-d"))
    return inspections, inspections_summary_data, inspections_summary_labels


def _total_count(data):
    total_count = reduce(lambda x, y: x + y, [x["小計"] for x in data.values()])
    return total_count


def parse_inspection_per_date(filepath):
    data = _inspection_dataset_from_chiba_pref(filepath)
    data = _fill_data(data)
    total_count = _total_count(data)
    return _data_to_inspections(data), total_count


if __name__ == '__main__':
    filename = "chiba.xlsx"
    filepath = os.path.join(*[os.path.abspath(os.path.dirname(__file__)), '../data', filename])
    print(parse_inspection_per_date(filepath))
