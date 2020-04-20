"""
コールセンター相談件数-RAW.xlsx
"""

from openpyxl import load_workbook
import os
import sys
from pathlib import Path
sys.path.append(str(Path('__file__').resolve().parent))
from common import excel_date

FILENAME = "コールセンター相談件数-RAW.xlsx"

def call_center_modified():
    paths = [os.path.abspath(os.path.dirname(__file__)), '..', 'data', FILENAME]
    f = os.path.join(*paths)
    wb = load_workbook(f)
    return wb.properties.modified

def parse_call_center():
    paths = [os.path.abspath(os.path.dirname(__file__)), '..', 'data', FILENAME]
    f = os.path.join(*paths)
    wb = load_workbook(f)
    ws = wb.active
    results = []
    count = 0
    for row in ws.values:
        count += 1
        if count == 1:
            continue
        if not row[0]:
            break
        date = excel_date(row[0])
        call_in_morning = int(row[2] or "0")
        call_in_afternoon = int(row[3] or "0")
        call_in_night = int(row[4] or "0")
        sum_call_count = call_in_morning + call_in_afternoon + call_in_night
        results.append({
            "日付": date.isoformat(timespec='milliseconds')+'Z',
            "曜日": row[1],
            "9-13時": call_in_morning,
            "13-17時": call_in_afternoon,
            "17-21時": call_in_night,
            "date": date.strftime('%Y-%m-%d'),
            "w": date.strftime("%w"),
            "short_date": date.strftime("%m-%d"),
            "小計": sum_call_count
        })
    return results

if __name__ == '__main__':
    print(parse_call_center())
