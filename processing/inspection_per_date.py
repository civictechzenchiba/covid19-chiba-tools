"""
- 検査実施日別状況.xlsx
- 検査実績（データセット）千葉県衛生研究所2019-nCoVラインリスト<日付>.xlsx

上記を集計したものを出力

詳細なデータ
```
{
    "判明日": "1\/25\/2020",
    "検査検体数": "3 ",
    "疑い例検査": "3 ",
    "接触者調査": "0 ",
    "陰性確認": "0 ",
    "（小計①）": "3 ",
    "チャーター便・クルーズ便等": "0 ", # 2020/3/22データ形式変更
    "陰性確認2": "0 ",
    "（小計②）": "0 "
},
```

"""

from openpyxl import load_workbook
import os
import glob
from datetime import timedelta
from functools import reduce
from pathlib import Path
import sys
sys.path.append(str(Path('__file__').resolve().parent))
from common import excel_date

def _empty_data(date):
    return {
        "判明日": date.strftime('%-m/%-d/%Y'),
        "検査検体数": 0,
        "疑い例検査": 0,
        "接触者調査": 0,
        "陰性確認": 0,
        "（小計①）": 0,
        "チャーター便・クルーズ便等": 0,
        "陰性確認2": 0,
        "（小計②）": 0
    }

def inspections_modified():
    FILENAME = "検査実績*千葉県衛生研究所2019-nCoVラインリスト*.xlsx"
    paths = [os.path.abspath(os.path.dirname(__file__)), '..', 'data', FILENAME]
    f = glob.glob(os.path.join(*paths))[0]
    wb = load_workbook(f)
    return wb.properties.modified

def _inspection_dataset_from_chiba_pref():
    FILENAME = "検査実績*千葉県衛生研究所2019-nCoVラインリスト*.xlsx"
    """
    検査実績（データセット）の処理
    """
    paths = [os.path.abspath(os.path.dirname(__file__)), '..', 'data', FILENAME]
    f = glob.glob(os.path.join(*paths))[0]
    data = {}
    wb = load_workbook(f)
    ws = wb.active
    i = 0
    for row in ws.values:
        i += 1
        if i == 1: # pass header
            continue
        if not row[0]: # pass empty row
            continue
        definite_date = row[8]
        target_date = definite_date.date()
        if not target_date in data.keys():
            data[target_date] = _empty_data(target_date)
        sample_count = row[9]
        if sample_count:
            data[target_date]["検査検体数"] += sample_count
        category = row[3]
        negative = row[10] == "陰性"
        if category == "一般":
            data[target_date]["疑い例検査"] += 1
            if negative:
                data[target_date]["陰性確認"] += 1
        elif category == "濃厚接触者":
            data[target_date]["接触者調査"] += 1
            if negative:
                data[target_date]["陰性確認"] += 1
        elif category == "チャーター" or category == "クルーズ" or category == "空港検疫":
            data[target_date]["チャーター便・クルーズ便等"] += 1
            if negative:
                data[target_date]["陰性確認2"] += 1
    return data

def _inspection_list_from_chiba_city(data):
    FILENAME = "検査実施日別状況.xlsx"
    paths = [os.path.abspath(os.path.dirname(__file__)), '..', 'data', FILENAME]
    f = os.path.join(*paths)
    wb = load_workbook(f)
    ws = wb.active
    i = 0
    for row in ws.values:
        i += 1
        if i == 1: # pass header
            continue
        if not row[0]: #pass empty row
            continue
        definite_date = excel_date(row[0])
        target_date = definite_date.date()
        if not target_date in data.keys():
            data[target_date] = _empty_data(target_date)
        data[target_date]["検査検体数"] += int(row[1] or '0') # 今の所0件になってしまう
        data[target_date]["疑い例検査"] += int(row[2] or '0')
        data[target_date]["接触者調査"] += int(row[3] or '0')
        data[target_date]["陰性確認"] += int(row[4] or '0') # 今の所0件になってしまう
        # 小計はあとでカウント
        data[target_date]["チャーター便・クルーズ便等"] += int(row[6] or '0')
        data[target_date]["チャーター便・クルーズ便等"] += int(row[7] or '0')
        data[target_date]["陰性確認2"] += int(row[8] or '0') # 今の所0件になってしまう
        # 小計はあとでカウント
    return data

def _sum_data(data):
    for target_date in data.keys():
        pt1 = data[target_date]["疑い例検査"]
        pt2 = data[target_date]["接触者調査"]
        data[target_date]["（小計①）"] = pt1 + pt2
        pt3 = data[target_date]["チャーター便・クルーズ便等"]
        data[target_date]["（小計②）"] = pt3
    return data

def _fill_data(data):
    """
    空の日付にデータを埋める
    """
    from_day = min(data.keys())
    to_day = max(data.keys())
    for i in range((to_day - from_day).days + 1):
        d = from_day + timedelta(i)
        if not d in data.keys():
            data[d] = _empty_data(d)
    return data

def _data_to_inspections(data):
    """
    inspectionsのデータを作成
    """
    inspections = []
    inspections_summary_data = {
        "県内": [],
        "その他": []
    }
    inspections_summary_labels = []
    keys = sorted(data.keys())
    for key in keys:
        inspections.append(data[key])
        inspections_summary_data["県内"].append(data[key]["（小計①）"])
        inspections_summary_data["その他"].append(data[key]["（小計②）"])
        inspections_summary_labels.append(key.strftime("%-m/%-d"))
    return inspections, inspections_summary_data, inspections_summary_labels

def _total_count(data):
    total_count = reduce(lambda x, y: x + y, [x["（小計①）"] for x in data.values()])
    total_count += reduce(lambda x, y: x + y, [x["（小計②）"] for x in data.values()])
    return total_count

def parse_inspection_per_date():
    data = _inspection_dataset_from_chiba_pref()
    data = _inspection_list_from_chiba_city(data)
    data = _sum_data(data)
    data = _fill_data(data)
    total_count = _total_count(data)
    return  _data_to_inspections(data), total_count

if __name__ == '__main__':
    print(parse_inspection_per_date())


